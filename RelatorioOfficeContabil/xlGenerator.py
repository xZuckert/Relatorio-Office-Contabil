from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

def gerarExcel(grupos, nome_saida="Relatorio.xlsx"):
    # cria a planilha na memoria seleciona a planilha ativa e renomeia
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    # cria o cabeçalho da planilha
    ws.append(["Dia", "Série", "número", "valor contabil (R$)"])

    totalGeral = 0 # variavel da soma total de todos os dias

    # Percorre as chaves do dicionario recupera os devidos valores e adiciona a linha na planilha
    for (dia, serie) in sorted(grupos):
        dados = grupos[(dia, serie)]
        intervalo = f"{dados['min']}-{dados['max']}"
        valor = round(dados["total"], 2)

        totalGeral += valor
        ws.append([dia, serie, intervalo, valor])

    ws.append(["", "", "TOTAL", round(totalGeral, 2)]) # Adiciona o valor total do mes

    #Estilo da planilha
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
        c.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.border = border
            if c.column == 4:
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="center")

    ultima_linha = ws.max_row
    for c in ws[ultima_linha]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FFF2CC")

    tabela = Table(displayName="TabelaRelatorio", ref=f"A1:D{ultima_linha-1}")
    estilo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)

    wb.save(nome_saida)
    print(f"Planilha gerada com sucesso: {nome_saida}")